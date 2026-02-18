import cv2
import os
import numpy as np
import openpyxl
from PIL import Image, ImageTk
from flask import Flask, render_template, request, jsonify

app = Flask(__name__)
@app.route("/home")
def home():
    message = "Welcome to Face Recognition System"
    return render_template('index.html', message=message)
# Function to capture faces from webcam with error handling
@app.route('/capture', methods=['POST'])
def capture_faces():
    # Initialize the webcam
    try:
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            raise Exception("Failed to open the webcam")
    except Exception as e:
        return jsonify({"error": "Error opening webcam: " + str(e)}), 500
    face_cascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
    # Read input text from the request
    name = request.form.get('name')
    # Load the Excel file
    path = "D:/Face Recognition/detail.xlsx"
    try:
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        max_row = sheet_obj.max_row
        cell_obj = sheet_obj.cell(row=max_row, column=1)
        id = cell_obj.value + 1
    except Exception as e:
        return jsonify({"error": f"Error loading Excel file: {str(e)}"}), 500
    # Capture faces
    img_id = 0
    while True:
        ret, my_frame = cap.read()
        if not ret:
            break
        gray = cv2.cvtColor(my_frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)
        for (x, y, w, h) in faces:
            face_cropped = my_frame[y:y+h, x:x+w]
            if face_cropped is not None:
                img_id += 1
                face = cv2.cvtColor(face_cropped, cv2.COLOR_BGR2RGB)
                file_name_path = f"data/user.{id}.{img_id}.jpg"
                try:
                    cv2.imwrite(file_name_path, face)
                except Exception as e:
                    print(f"Error saving image: {str(e)}")  # Log the error
        if img_id == 10:
            break
    # Update Excel file with new ID and name (with error handling)
    try:
        sheet_obj.append([id, name])
        wb_obj.save(path)
    except Exception as e:
        print(f"Error updating Excel file: {str(e)}")  # Log the error
    # Release the webcam and return success response
    cap.release()
    message = "Faces captured successfully"
    return render_template('index.html', message=message)
# Function to train the face recognition model with improved data augmentation (consider adding error handling for file operations)
@app.route('/train', methods=['GET'])
def train_model():
    # Load face images and IDs
    data_dir = "data"
    path = [os.path.join(data_dir, file) for file in os.listdir(data_dir)]
    faces, labels = [], []
    for image_path in path:
        img = cv2.imread(image_path)
        if img is None:
            print(f"Error loading image: {image_path}")  # Log the error
            continue  # Skip problematic images
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        id = int(os.path.split(image_path)[1].split('.')[1])
        # Apply data augmentation techniques
        augmented_images = [gray]
        augmented_images.extend(perform_data_augmentation(gray))

        for augmented_image in augmented_images:
            faces.append(augmented_image)
            labels.append(id)

    # Convert labels to NumPy array
    labels = np.array(labels)
    # Train the face recognition model
    clf = cv2.face.LBPHFaceRecognizer_create()
    try:
        clf.train(faces, labels)
        clf.save("classifier.xml")
        message = "Model trained successfully"
    except Exception as e:
        message = f"Error training model: {str(e)}"
    return render_template('index.html', message=message)

# Function to perform data augmentation
def perform_data_augmentation(image):
    augmented_images = []
    # Flip horizontally
    flipped_image = cv2.flip(image, 1)
    augmented_images.append(flipped_image)
    # Adjust brightness and contrast
    for factor in np.linspace(0.5, 1.5, 3):
        augmented_image = cv2.convertScaleAbs(image, alpha=factor, beta=0)
        augmented_images.append(augmented_image)
    return augmented_images


@app.route('/recognize', methods=['GET'])
def recognize_faces():
    def draw(img, classifier, scaleFactor, minNeighbors, color, clf):
        gray_image = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        features = classifier.detectMultiScale(gray_image, scaleFactor, minNeighbors)
        for (x, y, w, h) in features:
            id, pred = clf.predict(gray_image[y:y + h, x:x + w])
            name = get_name_from_id(id)
            print("Confidence ", pred,name)
            # Handle unknown faces and low-confidence predictions
            if pred < 45:  
                cv2.rectangle(img, (x, y), (x + w, y + h), color, 2)
                cv2.putText(img, "Name: {}".format(name), (x + 5, y - 5), cv2.FONT_HERSHEY_COMPLEX, 0.5, (255, 255, 255), 2)
            else:
                cv2.rectangle(img, (x, y), (x + w, y + h), color, 2)
                cv2.putText(img, "Unknown", (x + 5, y - 5), cv2.FONT_HERSHEY_COMPLEX, 0.5, (255, 255, 255), 2)
        return img
    faceCascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
    clf = cv2.face.LBPHFaceRecognizer_create()
    try:
        clf.read("classifier.xml")
    except Exception as e:
        message = f"Error loading model: {str(e)}"
        return render_template('index.html', message=message)
    video_cap = cv2.VideoCapture(0)
    while True:
        ret, img = video_cap.read()
        img = draw(img, faceCascade, 1.1, 10, (0, 255, 0), clf)
        cv2.imshow("Welcome to face Recognition", img)
        if cv2.waitKey(1) == 27:
            break
    video_cap.release()
    cv2.destroyAllWindows()
    message = "Model Recognized successfully"
    return render_template('index.html', message=message)
def get_name_from_id(id):
    # Load Excel file and fetch name corresponding to the ID
    path = "D:/Face Recognition/detail.xlsx"
    try:
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        for row in sheet_obj.iter_rows(min_row=1, max_row=sheet_obj.max_row, min_col=1, max_col=2):
            if row[0].value == id:
                return row[1].value
        # If ID not found, return a specific message
        return "Unknown Person"
    except Exception as e:
        print(f"Error accessing Excel file: {str(e)}")  # Log the error
        return "Error: Unable to access data"
if __name__ == "__main__":
    app.run(debug=True)
